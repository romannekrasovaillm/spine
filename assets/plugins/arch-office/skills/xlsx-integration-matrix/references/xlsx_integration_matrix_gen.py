#!/usr/bin/env python3
"""Скелет генератора матрицы интеграций (xlsx, openpyxl).

Реестр стыков + сводка по системам (формулы COUNTIF по реестру).
Запуск: python3 xlsx_integration_matrix_gen.py matrix.xlsx
"""

import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

HEADERS = [
    "ID", "Источник", "Приёмник", "Поток", "Протокол/формат",
    "Частота/объём", "SLA", "Критичность", "Контракт", "Владелец", "Комментарий",
]
# ── ЗАПОЛНИТЬ ────────────────────────────────────────────────────────────────
LINKS = [
    ["INT-001", "Мобильный банк", "Платёжный хаб", "команда списания", "REST/JSON",
     "до 40 RPS", "p95 ≤ 300 мс", "критичный", "IS-12 v1.3", "payments-core", ""],
    ["INT-002", "Платёжный хаб", "Шина событий", "PaymentStatus", "Kafka/Avro",
     "~2 млн соб/сут", "доставка ≤ 5 с", "важный", "нет", "payments-core", "контракт оформить!"],
]
SYSTEMS = sorted({link[1] for link in LINKS} | {link[2] for link in LINKS})
# ─────────────────────────────────────────────────────────────────────────────


def style_header(ws, ncols: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"
    for j in range(1, ncols + 1):
        cell = ws.cell(row=1, column=j)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN


def main(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Стыки"
    for j, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=j, value=h)
    style_header(ws, len(HEADERS))
    widths = [9, 18, 18, 24, 14, 16, 16, 12, 12, 16, 24]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for row in LINKS:
        ws.append(row)
    for r in ws.iter_rows(min_row=2, max_row=1 + len(LINKS), max_col=len(HEADERS)):
        for cell in r:
            cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Сводка по системам: формулы по реестру (не ручные числа).
    sm = wb.create_sheet("Сводка по системам")
    sm_headers = ["Система", "Исходящих", "Входящих", "Без контракта", "Критичных"]
    for j, h in enumerate(sm_headers, start=1):
        sm.cell(row=1, column=j, value=h)
    style_header(sm, len(sm_headers))
    for j, w in enumerate([24, 12, 12, 14, 12], start=1):
        sm.column_dimensions[get_column_letter(j)].width = w
    last = 1 + len(LINKS)
    for i, system in enumerate(SYSTEMS, start=2):
        sm.cell(row=i, column=1, value=system)
        sm.cell(row=i, column=2, value=f'=COUNTIF(Стыки!B2:B{last},A{i})')
        sm.cell(row=i, column=3, value=f'=COUNTIF(Стыки!C2:C{last},A{i})')
        sm.cell(
            row=i, column=4,
            value=f'=COUNTIFS(Стыки!B2:B{last},A{i},Стыки!I2:I{last},"нет")'
                  f'+COUNTIFS(Стыки!C2:C{last},A{i},Стыки!I2:I{last},"нет")',
        )
        sm.cell(
            row=i, column=5,
            value=f'=COUNTIFS(Стыки!B2:B{last},A{i},Стыки!H2:H{last},"критичный")'
                  f'+COUNTIFS(Стыки!C2:C{last},A{i},Стыки!H2:H{last},"критичный")',
        )

    wb.save(path)
    print(f"записан {path}: стыков — {len(LINKS)}, систем в сводке — {len(SYSTEMS)}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "matrix.xlsx")
