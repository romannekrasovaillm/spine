#!/usr/bin/env python3
"""Скелет генератора матрицы решений (xlsx, openpyxl).

Итоги — формулы SUMPRODUCT (меняешь веса/оценки — лист пересчитывается).
Запуск: python3 xlsx_decision_matrix_gen.py decision.xlsx
"""

import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

# ── ЗАПОЛНИТЬ ────────────────────────────────────────────────────────────────
# Критерии: (название, вес %, определение/шкала)
CRITERIA = [
    ("Закрытие NFR", 30, "1 — не закрывает, 5 — закрывает штатно"),
    ("Стоимость владения", 25, "1 — дорого, 5 — дёшево (3 года, допущения в Обоснованиях)"),
    ("Зрелость/поддержка", 20, "1 — сырой, 5 — промышленный стандарт"),
    ("Обратимость решения", 15, "1 — необратимо, 5 — откат флагом"),
    ("Время внедрения", 10, "1 — > года, 5 — ≤ квартала"),
]
OPTIONS = ["Вариант A", "Вариант B", "0 — статус-кво"]
# Оценки: scores[критерий][вариант] (1–5)
SCORES = [
    [4, 5, 2],
    [3, 4, 5],
    [4, 3, 5],
    [2, 4, 5],
    [3, 4, 5],
]
# Обоснования: (критерий, вариант, текст/источник)
JUSTIFICATIONS = [
    ("Закрытие NFR", "Вариант B", "NFR-3 штатно: кластер из коробки (документация, 08.2026)"),
    ("Стоимость владения", "0 — статус-кво", "текущие расходы — отчёт ФО 2026, строка 14"),
]
# Сценарии чувствительности: (название, [веса по критериям])
SENSITIVITY = [
    ("Базовый", [30, 25, 20, 15, 10]),
    ("Все критерии равны", [20, 20, 20, 20, 20]),
    ("Риск важнее (обратимость 35%)", [25, 20, 15, 35, 5]),
]
# ─────────────────────────────────────────────────────────────────────────────


def style_header(ws, ncols: int) -> None:
    ws.freeze_panes = "A2"
    for j in range(1, ncols + 1):
        cell = ws.cell(row=1, column=j)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN


def main(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Матрица"
    nopts = len(OPTIONS)
    headers = ["Критерий", "Вес %", "Определение/шкала"] + OPTIONS
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    style_header(ws, len(headers))
    for j, w in enumerate([22, 8, 44] + [16] * nopts, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    first_data, last_data = 2, 1 + len(CRITERIA)
    for i, ((name, weight, definition), scores) in enumerate(zip(CRITERIA, SCORES), start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=weight)
        ws.cell(row=i, column=3, value=definition)
        for k, score in enumerate(scores):
            ws.cell(row=i, column=4 + k, value=score)
    total_row = last_data + 1
    ws.cell(row=total_row, column=1, value="Взвешенный итог")
    ws.cell(row=total_row, column=2, value=f"=SUM(B{first_data}:B{last_data})")
    for k in range(nopts):
        col = get_column_letter(4 + k)
        ws.cell(
            row=total_row, column=4 + k,
            value=f"=SUMPRODUCT($B{first_data}:$B{last_data},{col}{first_data}:{col}{last_data})/100",
        )
    rank_row = total_row + 1
    ws.cell(row=rank_row, column=1, value="Ранг")
    for k in range(nopts):
        col = get_column_letter(4 + k)
        first_col = get_column_letter(4)
        last_col = get_column_letter(3 + nopts)
        ws.cell(
            row=rank_row, column=4 + k,
            value=f"=RANK({col}{total_row},${first_col}${total_row}:${last_col}${total_row})",
        )
    for r in ws.iter_rows(min_row=total_row, max_row=rank_row, max_col=3 + nopts):
        for cell in r:
            cell.fill = TOTAL_FILL
            cell.font = Font(name="Calibri", size=11, bold=True)
    for r in ws.iter_rows(min_row=1, max_row=rank_row, max_col=3 + nopts):
        for cell in r:
            cell.border = THIN
            if cell.alignment is None or not cell.alignment.wrap_text:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    sn = wb.create_sheet("Чувствительность")
    sn_headers = ["Сценарий"] + OPTIONS
    for j, h in enumerate(sn_headers, start=1):
        sn.cell(row=1, column=j, value=h)
    style_header(sn, len(sn_headers))
    sn.column_dimensions["A"].width = 34
    for j in range(2, 2 + nopts):
        sn.column_dimensions[get_column_letter(j)].width = 16
    for i, (name, weights) in enumerate(SENSITIVITY, start=2):
        sn.cell(row=i, column=1, value=name)
        for k in range(nopts):
            terms = "+".join(
                f"{w}*Матрица!{get_column_letter(4 + k)}{2 + c}" for c, w in enumerate(weights)
            )
            sn.cell(row=i, column=2 + k, value=f"=({terms})/100")
    note_row = 2 + len(SENSITIVITY) + 1
    sn.cell(row=note_row, column=1, value="Вывод: при каких весах лидер меняется — заполнить по результатам сценариев.")

    js = wb.create_sheet("Обоснования")
    for j, h in enumerate(["Критерий", "Вариант", "Обоснование/источник"], start=1):
        js.cell(row=1, column=j, value=h)
    style_header(js, 3)
    for j, w in enumerate([22, 18, 80], start=1):
        js.column_dimensions[get_column_letter(j)].width = w
    for row in JUSTIFICATIONS:
        js.append(list(row))

    wb.save(path)
    print(f"записан {path}: критериев — {len(CRITERIA)}, вариантов — {nopts}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "decision.xlsx")
