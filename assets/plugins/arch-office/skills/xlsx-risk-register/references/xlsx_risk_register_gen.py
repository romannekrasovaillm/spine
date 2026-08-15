#!/usr/bin/env python3
"""Скелет генератора реестра архитектурных рисков (xlsx, openpyxl).

Листы: Реестр (баллы — формулы), Карта 5×5 (COUNTIFS + цветовые зоны), Легенда.
Запуск: python3 xlsx_risk_register_gen.py risks.xlsx
"""

import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
ZONE_GREEN = PatternFill("solid", fgColor="C6E0B4")
ZONE_YELLOW = PatternFill("solid", fgColor="FFE699")
ZONE_RED = PatternFill("solid", fgColor="F4B8C1")

HEADERS = [
    "ID", "Риск (причина → событие → следствие)", "Категория",
    "Вероятность", "Влияние", "Балл", "Митигация", "Владелец", "Срок",
    "Остат. вероятность", "Остат. влияние", "Остат. балл",
    "Статус", "Связи", "Дата пересмотра",
]
# ── ЗАПОЛНИТЬ ────────────────────────────────────────────────────────────────
RISKS = [
    ["R-01", "Из-за единой реплики хаба при сбое узла остановится приём платежей, что приведёт к простою каналов",
     "стойкость", 3, 5, None, "active-active в двух ЦОД + autoscale", "payments-core", "2026-Q4",
     1, 5, None, "mitigating", "F-1, ADR-014", "2026-10-01"],
    ["R-02", "Из-за отсутствия дедупликации повторная доставка статуса приведёт к двойному списанию",
     "данные", 4, 5, None, "идемпотентность по tx_id + outbox", "payments-core", "2026-Q3",
     1, 4, None, "open", "F-2", "2026-09-15"],
]
SCALES = {
    "Вероятность": ["1 — <5%/год", "2 — 5–20%", "3 — 20–50%", "4 — 50–80%", "5 — >80%"],
    "Влияние": ["1 — незаметно", "2 — локально", "3 — простой канала <1ч",
                "4 — простой >1ч / штраф", "5 — системный сбой / регуляторика"],
    "Статус": ["open", "mitigating", "accepted (ссылка на решение)", "closed"],
}
# ─────────────────────────────────────────────────────────────────────────────


def style_header(ws, ncols: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"
    for j in range(1, ncols + 1):
        cell = ws.cell(row=1, column=j)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN


def main(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр"
    for j, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=j, value=h)
    style_header(ws, len(HEADERS))
    widths = [7, 46, 12, 11, 9, 7, 34, 16, 10, 12, 11, 9, 11, 14, 12]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for i, row in enumerate(RISKS, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
        ws.cell(row=i, column=6, value=f"=D{i}*E{i}")     # исходный балл
        ws.cell(row=i, column=12, value=f"=J{i}*K{i}")    # остаточный балл
    for r in ws.iter_rows(min_row=2, max_row=1 + len(RISKS), max_col=len(HEADERS)):
        for cell in r:
            cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Карта 5×5: строки — влияние 5..1, колонки — вероятность 1..5.
    mp = wb.create_sheet("Карта 5×5")
    mp.cell(row=1, column=1, value="Влияние \\ Вероятность")
    for p in range(1, 6):
        mp.cell(row=1, column=1 + p, value=p)
    for i in range(5):
        impact = 5 - i
        mp.cell(row=2 + i, column=1, value=impact)
        for p in range(1, 6):
            score = impact * p
            cell = mp.cell(row=2 + i, column=1 + p)
            last = 1 + len(RISKS)
            cell.value = (
                f'=COUNTIFS(Реестр!D2:D{last},{p},Реестр!E2:E{last},{impact})'
            )
            cell.fill = ZONE_RED if score >= 15 else ZONE_YELLOW if score >= 6 else ZONE_GREEN
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN
    mp.column_dimensions["A"].width = 22
    for j in range(2, 7):
        mp.column_dimensions[get_column_letter(j)].width = 8
    mp.cell(row=8, column=1, value="Число рисков в клетке (исходная оценка). Красное ≥ 15, жёлтое ≥ 6.")

    lg = wb.create_sheet("Легенда")
    lg.append(["Шкала", "Определения"])
    for j in (1, 2):
        cell = lg.cell(row=1, column=j)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for name, values in SCALES.items():
        lg.append([name, "; ".join(values)])
    lg.column_dimensions["A"].width = 16
    lg.column_dimensions["B"].width = 90

    wb.save(path)
    print(f"записан {path}: рисков — {len(RISKS)}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "risks.xlsx")
