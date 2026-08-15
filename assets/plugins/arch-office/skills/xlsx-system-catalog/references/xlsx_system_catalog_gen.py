#!/usr/bin/env python3
"""Скелет генератора каталога систем (xlsx, openpyxl).

Запуск: python3 xlsx_system_catalog_gen.py catalog.xlsx
"""

import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

HEADERS = [
    "ID", "Название", "Домен", "Назначение", "Владелец (команда)",
    "Критичность", "Стек", "Жизненный цикл", "Инициативы/ADR", "Комментарий",
]
# ── ЗАПОЛНИТЬ ────────────────────────────────────────────────────────────────
SYSTEMS = [
    ["PAY-01", "Платёжный хаб", "Платежи", "Маршрутизация платежей между контурами",
     "payments-core", "критичная", "Java 21 / PostgreSQL / Kafka", "active", "ADR-014", ""],
    ["CHN-02", "Каналы: мобильный банк", "Каналы", "API мобильного клиента",
     "channels", "важная", "Kotlin / Ktor", "active", "", "планируется перенос в BFF"],
]
# Словари допустимых значений (лист «Легенда»).
LEGENDS = {
    "Критичность": ["критичная", "важная", "стандартная"],
    "Жизненный цикл": ["pilot", "active", "sunset", "retired"],
}
# ─────────────────────────────────────────────────────────────────────────────


def style_sheet(ws, widths: list[int]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    for j, (h, w) in enumerate(zip(HEADERS, widths), start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN
        ws.column_dimensions[get_column_letter(j)].width = w


def main(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Системы"
    style_sheet(ws, [9, 24, 14, 40, 20, 13, 30, 14, 14, 30])
    for row in SYSTEMS:
        ws.append(row)
    for r in ws.iter_rows(min_row=2, max_row=1 + len(SYSTEMS), max_col=len(HEADERS)):
        for cell in r:
            cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    legend = wb.create_sheet("Легенда")
    legend.append(["Словарь", "Допустимые значения"])
    for j in (1, 2):
        cell = legend.cell(row=1, column=j)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for name, values in LEGENDS.items():
        legend.append([name, ", ".join(values)])
    legend.column_dimensions["A"].width = 20
    legend.column_dimensions["B"].width = 60

    wb.save(path)
    print(f"записан {path}: систем — {len(SYSTEMS)}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "catalog.xlsx")
